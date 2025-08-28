import pandas as pd
from openpyxl import load_workbook
import random
import re
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import time
import os
import sys

def is_valid_url(url):
    """检查URL是否以http开头"""
    if not url:
        return False
    return str(url).strip().lower().startswith(('http://', 'https://'))

def get_hyperlinks_from_excel(excel_file):
    """从Excel文件第一列提取超链接的实际URL"""
    wb = load_workbook(excel_file, read_only=False, data_only=False)
    ws = wb.active  # 获取活动工作表
    
    hyperlinks = []
    hyperlink_pattern = re.compile(r'HYPERLINK\("(.*?)",')
    
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        url = None
        
        # 提取超链接
        if hasattr(cell, 'hyperlink') and cell.hyperlink:
            if hasattr(cell.hyperlink, 'target'):
                url = cell.hyperlink.target
            elif hasattr(cell.hyperlink, 'url'):
                url = cell.hyperlink.url
        
        # 从公式提取
        if not url and cell.value:
            cell_value = str(cell.value)
            match = hyperlink_pattern.search(cell_value)
            if match:
                url = match.group(1)
        
        # 直接识别URL格式
        if not url:
            cell_value = str(cell.value).strip() if cell.value else ""
            if cell_value.startswith(('http://', 'https://', 'www.')):
                url = cell_value
        
        hyperlinks.append(url if url else cell_value if cell_value else None)
    
    wb.close()
    return hyperlinks

def save_progress(df, excel_file):
    """保存当前进度到Excel文件"""
    try:
        df.to_excel(excel_file, index=False)
        print(f"\n已保存当前进度到 {excel_file}")
    except Exception as e:
        print(f"\n保存进度时出错: {str(e)}")

def find_introduce_element(page):
    """多种方式查找introduce元素"""
    # 方法1: 直接查找class
    locator = page.locator(".introduce")
    if locator.count() > 0:
        return locator
    
    # 方法2: 查找包含该class的任何标签（可能有嵌套）
    locator = page.locator("*[class*='introduce']")
    if locator.count() > 0:
        return locator
    
    # 方法3: 等待并尝试滚动到元素
    try:
        # 先滚动到页面底部触发可能的懒加载
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1)
        # 再滚动到顶部
        page.evaluate("window.scrollTo(0, 0)")
        time.sleep(1)
        
        # 再次尝试查找
        locator = page.locator(".introduce")
        if locator.count() > 0:
            # 滚动到元素可见
            locator.scroll_into_view_if_needed()
            time.sleep(0.5)
            return locator
    except:
        pass
    
    return None

def process_links_from_excel(excel_file):
    # 读取Excel文件
    df = pd.read_excel(excel_file)
    actual_urls = get_hyperlinks_from_excel(excel_file)
    actual_urls = [url for url in actual_urls if url is not None]
    
    if not actual_urls:
        print("Excel文件中没有有效的超链接")
        return
    
    # 让用户输入起始行（从1开始）
    try:
        start_row = input(f"请输入起始行（1-{len(df)}）: ").strip()
        start_row = int(start_row) - 1  # 转换为0-based索引
        if start_row < 0 or start_row >= len(df):
            print(f"无效的起始行，将从第1行开始")
            start_row = 0
    except ValueError:
        print("输入无效，将从第1行开始")
        start_row = 0
    
    # 显示提取到的URL供检查
    print("\n提取到的URL列表（前5个）:")
    for i in range(min(5, len(actual_urls))):
        print(f"{i+1}. {actual_urls[i]}")
    if len(actual_urls) > 5:
        print(f"... 还有 {len(actual_urls)-5} 个URL")
    
    # 确认URL是否正确
    confirm = input("\n这些URL看起来正确吗？(y/n): ").strip().lower()
    if confirm != 'y':
        print("请检查Excel文件中的超链接格式")
        return
    
    # 确保有第五列
    if df.shape[1] < 5:
        for i in range(df.shape[1], 5):
            df[f"Unnamed: {i}"] = ""
        df = df.rename(columns={df.columns[4]: "Content"})
    
    # 查找第一个有效的URL（用于登录）
    valid_index = 0
    while valid_index < len(actual_urls) and not is_valid_url(actual_urls[valid_index]):
        valid_index += 1
    if valid_index >= len(actual_urls):
        print("未找到有效的URL")
        return
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        
        try:
            # 处理第一个链接（登录用）
            first_url = actual_urls[valid_index]
            print(f"\n打开登录链接: {first_url}")
            page.goto(first_url)
            input("请在浏览器中完成登录，登录完成后按Enter继续...")
            
            # 如果起始行是0，处理第一个页面内容
            if start_row == 0:
                content = ""
                # 增加等待时间并使用增强的查找方法
                page.wait_for_load_state("load", timeout=30000)
                time.sleep(2)  # 额外等待2秒
                
                locator = find_introduce_element(page)
                if locator:
                    content = locator.inner_text()
                    print(f"\n第1行内容预览:\n{content[:200]}...\n")
                else:
                    print("第1行页面中未找到class为'introduce'的元素，尝试手动检查...")
                    # 给用户时间手动确认
                    input("请确认页面中是否有introduce元素，确认后按Enter继续...")
                    # 再次尝试
                    locator = find_introduce_element(page)
                    if locator:
                        content = locator.inner_text()
                        print(f"找到元素，内容预览:\n{content[:200]}...\n")
                
                df.iloc[0, 4] = content
                save_progress(df, excel_file)  # 保存第一行结果
            
            # 处理从起始行开始的链接
            for i in range(start_row, len(actual_urls)):
                # 跳过无效URL
                if not is_valid_url(actual_urls[i]):
                    print(f"\n第{i+1}行URL无效，跳过")
                    continue
                
                # 跳过已处理的行（如果第五列已有内容）
                if pd.notna(df.iloc[i, 4]) and str(df.iloc[i, 4]).strip() != "":
                    print(f"\n第{i+1}行已处理，跳过")
                    continue
                
                print(f"\n处理第{i+1}行链接: {actual_urls[i]}")
                new_page = context.new_page()
                content = ""
                try:
                    new_page.goto(actual_urls[i], timeout=30000)
                    new_page.wait_for_load_state("load", timeout=30000)
                    time.sleep(2)  # 额外等待2秒
                    
                    locator = find_introduce_element(new_page)
                    if locator:
                        content = locator.inner_text()
                        print(f"内容预览:\n{content[:200]}...\n")
                    else:
                        print("未找到class为'introduce'的元素")
                
                except PlaywrightTimeoutError:
                    print("页面加载超时")
                    content = "错误: 页面加载超时"
                except Exception as e:
                    print(f"处理时出错: {str(e)}")
                    content = f"错误: {str(e)}"
                finally:
                    df.iloc[i, 4] = content
                    time.sleep(random.uniform(1, 3))
                    new_page.close()
                    save_progress(df, excel_file)  # 每处理一行就保存一次
                
                time.sleep(random.uniform(10, 40))
            
            print("\n所有指定行处理完成")
        
        except Exception as e:
            print(f"\n发生错误: {str(e)}")
            save_progress(df, excel_file)  # 出错时保存当前进度
        finally:
            browser.close()
            print(f"\n最终结果已保存到 {excel_file}")

if __name__ == "__main__":
    process_links_from_excel("chan.xlsx")