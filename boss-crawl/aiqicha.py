import re
import time
import random
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

def extract_contact_info(text):
    """从文本中提取手机号和邮箱"""
    # 手机号正则（11位数字，以1开头，第二位3-9）
    phone_pattern = re.compile(r'1[3-9]\d{9}')
    # 邮箱正则（支持常见格式）
    email_pattern = re.compile(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+')
    
    phones = phone_pattern.findall(text)
    emails = email_pattern.findall(text)
    
    # 取第一个匹配结果，无则返回空
    phone = phones[0] if phones else ""
    email = emails[0] if emails else ""
    return phone, email

def main():
    # 加载Excel文件
    wb = load_workbook("BOSSid.xlsx")
    ws = wb.active  # 获取活动工作表
    max_row = ws.max_row  # 获取最大行数
    print(f"共检测到 {max_row - 1} 条公司数据（跳过表头）")
    start_row_index = 2

    with sync_playwright() as p:
        # 启动浏览器（无头模式关闭，方便手动登录）
        browser = p.chromium.launch(headless=False, slow_mo=50)
        page = browser.new_page()
        
        # 打开天眼查搜索页
        page.goto("https://www.tianyancha.com/nsearch?key=")
        print("请在60秒内手动完成登录（扫码/账号密码）...")
        time.sleep(60)  # 等待登录
        
        # 遍历公司名称（从第2行开始，跳过表头）
        for row in range(start_row_index, max_row + 1):
            company_name = ws.cell(row=row, column=3).value  # 第三列（索引2）
            if not company_name:
                print(f"第{row}行无公司名称，跳过")
                continue
            
            try:
                print(f"\n处理第{row}行：{company_name}")
                
                # 定位搜索框并输入公司名称（清空后输入，避免残留内容）
                search_input = page.wait_for_selector(
                    ".tyc-header-suggest-content input",
                    timeout=10000
                )
                search_input.fill('')
                search_input.fill(company_name)
                time.sleep(random.uniform(0.5, 1.5))  # 模拟人工输入间隔
                
                # 点击搜索按钮
                search_btn = page.wait_for_selector(
                    ".tyc-header-suggest-button",
                    timeout=10000
                )
                search_btn.click()
                time.sleep(random.uniform(2, 4))  # 等待结果加载
                
                # 提取第一个结果的联系信息
                contact_element = page.wait_for_selector(
                    ".index_contact-row__iYUn6",
                    timeout=15000  # 最长等待15秒
                )
                contact_text = contact_element.text_content() or ""
                print(f"提取到联系文本：{contact_text[:50]}...")  # 打印前50字符
                
                # 解析手机号和邮箱
                phone, email = extract_contact_info(contact_text)
                print(f"匹配结果：手机号={phone}，邮箱={email}")
                
                # 写入Excel（第六列=手机号，第七列=邮箱）
                ws.cell(row=row, column=6).value = phone
                ws.cell(row=row, column=7).value = email
                wb.save("BOSSid.xlsx")  # 实时保存
                
            except Exception as e:
                print(f"第{row}行处理失败：{str(e)}")
                # 失败时写入空值
                ws.cell(row=row, column=6).value = ""
                ws.cell(row=row, column=7).value = ""
                wb.save("BOSSid.xlsx")
            
            # 随机等待，降低反爬风险
            time.sleep(random.uniform(3, 6))
    
    print("\n所有公司处理完成，结果已写入BOSSid.xlsx")
    wb.close()

if __name__ == "__main__":
    main()